"""
Excel 파일 파서 — parts / part_deals 데이터 추출

지원 형식:
  1. 비정규화 형식 (한 시트에 부품+납품 혼합):
     - 헤더 자동 탐색 (1~15행)
     - 헤더 기반 매핑 + 위치 기반 폴백
  2. 정규화된 시트 형식 (부품 시트 + 납품 시트 분리)

위치 기반 기본 컬럼 (헤더가 없거나 매칭 실패 시):
  A=시리즈, B=직품모델, C=품목코드, D=품명, E=거래수,
  F=원가, G=로컬가, H=종병가, I=교체증상, J=상세설명, K=교체위치,
  L=병원명, M=납품일, N=수량, O=납품단가, P=원가(납품)
"""
from __future__ import annotations
import io
from datetime import date, datetime
from openpyxl import load_workbook
from app.schemas.part import SeedData, SeedRow, SeedDealRow


# ── 회사/카테고리 분류 규칙 ──────────────────────────────────────────
SERIES_RULES: list[tuple[str, str, list[str]]] = [
    # KOWA (Japan)
    ("KOWA (Japan)", "Fundus Camera (NONMYD 시리즈)", ["nonmyd"]),
    ("KOWA (Japan)", "Fundus Camera (VX 시리즈)",    ["vx-10", "vx-20", "vx 10", "vx series", "fundus camera (vx"]),
    ("KOWA (Japan)", "Fundus Camera (GENESIS 시리즈)", ["genesis"]),
    ("KOWA (Japan)", "Fundus Camera",               ["fundus"]),
    ("KOWA (Japan)", "Slit Lamp (세극등)",           ["slit lamp", "세극등"]),
    ("KOWA (Japan)", "Tonometer (안압계)",           ["tonometer", "안압계"]),
    ("KOWA (Japan)", "Perimeter (시야계)",           ["perimeter", "시야계"]),
    ("KOWA (Japan)", "Indirect Ophthalmoscope (도상검안경)", ["indirect ophthalmoscope", "도상검안경"]),
    ("KOWA (Japan)", "Digital Conversion (디지털 변환)", ["digital conversion", "디지털 변환"]),
    # A.R.C. Laser (Germany)
    ("A.R.C. Laser (Germany)", "Laser Fiber/Probe (파이버/프로브)",
     ["lipolysis", "cyclophoto", "endo probe", "laser fiber", "laser probe", "bare fiber", "fiber"]),
    ("A.R.C. Laser (Germany)", "Laser (레이저)",
     ["q-las", "fox q", "classic (laser)", "classic laser", "safety filter", "safety goggle",
      "goggle", "foot switch q", "laser table", "laser filter", "laser"]),
    # KONAN (Japan)
    ("KONAN (Japan)", "Specular Microscope", ["konan", "specular", "nsp-", "nsp ", "cellchek", "sp-"]),
    # KEELER (UK)
    ("KEELER (UK)", "Loupe",                 ["loupe"]),
    ("KEELER (UK)", "Slit Lamp",             ["keeler slit", "psl-"]),
    ("KEELER (UK)", "Indirect Ophthalmoscope", ["vantageplus", "api ii", "symphony", "indirect"]),
    ("KEELER (UK)", "Accessories",           ["keeler", "handle", "charger", "battery"]),
    # Phaco Handpiece
    ("Phaco Handpiece (해외)", "Handpiece",
     ["handpiece", "phaco", "infiniti", "centurion", "sovereign", "legacy", "millennium", "stellaris"]),
    # Camera/주변기기
    ("Camera/주변기기", "Camera",            ["ccd camera", "camera"]),
    ("Camera/주변기기", "Accessories",       ["monitor", "printer", "video"]),
    # NewEyesTech
    ("NewEyesTech", "Digital Adapters",     ["beam splitter", "video adapter", "camera adapter", "neweyes"]),
    ("NewEyesTech", "Digital Conversion",   ["digital"]),
]


def _classify_series(series: str) -> tuple[str, str]:
    low = (series or "").lower()
    for company, category, keywords in SERIES_RULES:
        if any(kw in low for kw in keywords):
            return company, category
    return "기타", "기타"


# ── 유틸리티 ────────────────────────────────────────────────────────
def _clean(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v != v:  # NaN
        return ""
    return str(v).strip()


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        s = str(v).replace(",", "").replace("원", "").replace("만", "").strip()
        if not s:
            return None
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y.%m.%d", "%Y/%m/%d", "%m/%d/%Y", "%Y년 %m월 %d일",
                "%Y-%m-%d %H:%M:%S", "%Y.%m", "%Y/%m"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


# ── 헤더 키워드 매핑 (lower-case 부분 일치) ─────────────────────────
# 정확 일치(EXACT)와 부분 일치(KEYWORDS)를 분리해서 처리.
# 정확 일치가 우선 — 영문 컬럼명("company", "model" 등)이 잘못 매칭되는 것 방지.
HEADER_EXACT: dict[str, str] = {
    # 부품 마스터 (영문 정확 매칭)
    "id": "_skip",
    "company": "company",
    "category": "category",
    "model": "model",
    "model_name": "model",
    "part_code": "part_code",
    "part code": "part_code",
    "partcode": "part_code",
    "part_name": "part_name",
    "part name": "part_name",
    "partname": "part_name",
    "cost_avg": "cost_avg",
    "cost": "cost_avg",
    "local_price": "local_price",
    "local": "local_price",
    "univ_price": "univ_price",
    "univ price": "univ_price",
    "deal_count": "deal_count",
    "deal count": "deal_count",
    "avg_price": "avg_price",
    "min_price": "min_price",
    "max_price": "max_price",
    "symptom": "symptom",
    "symptom_detail": "symptom_detail",
    "detail": "symptom_detail",
    "symptom_location": "symptom_location",
    "location": "symptom_location",
    # 납품
    "part_id": "_skip",
    "hospital": "hospital",
    "deal_date": "deal_date",
    "deal date": "deal_date",
    "quantity": "quantity",
    "qty": "quantity",
    "deal_price": "deal_price",
    "deal price": "deal_price",
    "cost_price": "cost_price",
    "cost price": "cost_price",
    # 한글 정확 매칭
    "회사": "company",
    "제조사": "company",
    "분류": "category",
    "카테고리": "category",
    "형명": "category",
    "모델명": "model",
    "모델": "model",
    "시리즈": "model",
    "series": "model",
    "키워드": "_skip",
}

# 부분 일치 (한글 위주 — 영문 충돌 없도록)
HEADER_KEYWORDS: list[tuple[list[str], str]] = [
    (["품목코드", "품번", "파트번호", "품목 코드"], "part_code"),
    (["품명", "부품명"], "part_name"),
    (["직품모델", "제품모델"], "model2"),
    (["거래수", "납품건수", "거래건수"], "deal_count"),
    (["로컬"], "local_price"),
    (["종병", "대학", "종합병원"], "univ_price"),
    (["교체 증상", "교체증상", "증상"], "symptom"),
    (["상세 설명", "상세설명", "설명"], "symptom_detail"),
    (["교체 위치", "교체위치", "위치"], "symptom_location"),
    (["병원명", "병원", "거래처"], "hospital"),
    (["납품일", "거래일"], "deal_date"),
    (["수량"], "quantity"),
    (["납품단가", "납품가", "거래단가", "거래가"], "deal_price"),
    (["실원가", "납품원가"], "cost_price"),
    (["원가"], "cost_avg"),  # cost_price보다 뒤 — 실원가가 먼저 매칭되도록
]


def _norm_cell(v) -> str:
    return _clean(v).lower().replace("\n", " ").replace("  ", " ")


def _try_match_field(cell_text: str) -> str | None:
    """헤더 셀 텍스트 → 필드명 (없으면 None). 정확 일치 → 부분 일치 순."""
    if not cell_text:
        return None
    # 정확 일치 우선
    if cell_text in HEADER_EXACT:
        f = HEADER_EXACT[cell_text]
        return None if f == "_skip" else f
    # 부분 일치
    for keywords, field in HEADER_KEYWORDS:
        for kw in keywords:
            if kw in cell_text:
                return field
    return None


def _find_header_row(rows: list[tuple]) -> tuple[int, dict[str, int]] | None:
    """
    rows 중에서 헤더처럼 보이는 행을 찾아 (행 인덱스, {field: col_idx}) 반환.
    찾는 기준: part_code 또는 hospital + 1개 이상 다른 필드를 가진 행.
    """
    best: tuple[int, dict[str, int]] | None = None
    for i, row in enumerate(rows[:20]):  # 첫 20행만 검사
        col: dict[str, int] = {}
        for j, cell in enumerate(row):
            text = _norm_cell(cell)
            field = _try_match_field(text)
            if field and field not in col:
                col[field] = j
        # 헤더로 인정: part_code 또는 hospital 같은 핵심 필드 + 다른 필드 1개 이상
        has_key = "part_code" in col or "hospital" in col or "part_name" in col
        if has_key and len(col) >= 2:
            if best is None or len(col) > len(best[1]):
                best = (i, col)
    return best


# ── 위치 기반 기본 컬럼 ──────────────────────────────────────────────
# 사용자 사양에 따른 기본 위치 (0-indexed)
DEFAULT_POSITIONS = {
    "series":     0,   # A
    "model2":     1,   # B
    "part_code":  2,   # C
    "part_name":  3,   # D
    "deal_count": 4,   # E
    "cost_avg":   5,   # F
    "local_price": 6,  # G
    "univ_price": 7,   # H
    "symptom":    8,   # I
    "symptom_detail": 9,  # J
    "symptom_location": 10,  # K
    "hospital":   11,  # L
    "deal_date":  12,  # M
    "quantity":   13,  # N
    "deal_price": 14,  # O
    "cost_price": 15,  # P
}


def _make_getter(row, col_idx: dict[str, int], use_positions: bool = False):
    """필드명 → 셀 값. col_idx 우선, 없으면 DEFAULT_POSITIONS 폴백."""
    def get(field):
        idx = col_idx.get(field)
        if idx is None and use_positions:
            idx = DEFAULT_POSITIONS.get(field)
        if idx is None or idx >= len(row):
            return None
        return row[idx]
    return get


# ── 비정규화 형식 파서 ────────────────────────────────────────────────
def _parse_denormalized(
    sheet,
    company_override: str | None = None,
) -> tuple[list[SeedRow], list[SeedDealRow]]:
    all_rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    if not all_rows:
        return [], []

    # 1) 헤더 행 탐색
    header_match = _find_header_row(all_rows)
    if header_match:
        header_idx, col_idx = header_match
        data_rows = all_rows[header_idx + 1:]
    else:
        # 헤더 못 찾으면 모든 행을 데이터로 간주 (위치 기반).
        # 부품 마스터 행 등록 시 part_code가 헤더 단어면 자동 스킵됨.
        col_idx = {}
        data_rows = all_rows
    use_positions = True

    # 2) 데이터 파싱
    parts: dict[str, SeedRow] = {}
    deals: list[SeedDealRow] = []
    current_code: str | None = None
    current_series: str | None = None

    for row in data_rows:
        if not row or all(v is None or _clean(v) == "" for v in row):
            continue

        get = _make_getter(row, col_idx, use_positions=use_positions)

        series = _clean(get("series"))
        part_code = _clean(get("part_code"))
        part_name = _clean(get("part_name"))
        hospital = _clean(get("hospital"))
        deal_date_raw = get("deal_date")
        deal_price_raw = get("deal_price")

        # part_code가 너무 짧거나 헤더 단어면 스킵
        if part_code and (
            part_code.lower() in ("품목코드", "part_code", "part code", "품번")
        ):
            continue

        # ── 부품 마스터 행 ──
        if part_code and part_code not in parts:
            # 명시적 company/category/model 컬럼이 있으면 우선
            explicit_company = _clean(get("company"))
            explicit_category = _clean(get("category"))
            explicit_model = _clean(get("model"))
            if not explicit_company or not explicit_category:
                cls_c, cls_cat = _classify_series(explicit_model or series)
                company = explicit_company or cls_c
                category = explicit_category or cls_cat
            else:
                company, category = explicit_company, explicit_category
            if company_override:
                company = company_override

            parts[part_code] = SeedRow(
                company=company,
                category=category,
                model=explicit_model or series or "기타",
                part_code=part_code,
                part_name=part_name or part_code,
                cost_avg=_to_int(get("cost_avg")),
                local_price=_to_int(get("local_price")),
                univ_price=_to_int(get("univ_price")),
                symptom=_clean(get("symptom")) or None,
                symptom_detail=_clean(get("symptom_detail")) or None,
                symptom_location=_clean(get("symptom_location")) or None,
                deal_count=_to_int(get("deal_count")) or 0,
                avg_price=_to_int(get("avg_price")),
                min_price=_to_int(get("min_price")),
                max_price=_to_int(get("max_price")),
            )
            current_code = part_code
            current_series = series

        elif part_code and part_code in parts:
            current_code = part_code
            current_series = series

        elif series and not part_code:
            if series != current_series:
                current_code = None
                current_series = series

        # ── 납품 내역 행 ──
        if current_code and (hospital or deal_date_raw or deal_price_raw):
            cost_price_raw = get("cost_price")
            if cost_price_raw is None:
                cost_price_raw = get("cost_avg")

            deals.append(SeedDealRow(
                part_code=current_code,
                hospital=hospital or None,
                deal_date=_to_date(deal_date_raw),
                quantity=_to_int(get("quantity")) or 1,
                deal_price=_to_int(deal_price_raw),
                cost_price=_to_int(cost_price_raw),
            ))

    return list(parts.values()), deals


# ── 정규화 시트 파서 (부품 시트) ─────────────────────────────────────
def _parse_parts_sheet(sheet) -> list[SeedRow]:
    all_rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    if not all_rows:
        return []
    header_match = _find_header_row(all_rows)
    if not header_match:
        return []
    header_idx, col_idx = header_match
    parts = []
    for row in all_rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        get = _make_getter(row, col_idx, use_positions=False)
        part_code = _clean(get("part_code"))
        part_name = _clean(get("part_name"))
        if not part_code or not part_name:
            continue
        # 명시적 company/category/model 컬럼 우선, 없으면 series로 분류
        company = _clean(get("company"))
        category = _clean(get("category"))
        model = _clean(get("model"))
        if not company or not category:
            cls_company, cls_category = _classify_series(model or _clean(get("series")))
            company = company or cls_company
            category = category or cls_category
        parts.append(SeedRow(
            company=company,
            category=category,
            model=model or "기타",
            part_code=part_code,
            part_name=part_name,
            cost_avg=_to_int(get("cost_avg")),
            local_price=_to_int(get("local_price")),
            univ_price=_to_int(get("univ_price")),
            symptom=_clean(get("symptom")) or None,
            symptom_detail=_clean(get("symptom_detail")) or None,
            symptom_location=_clean(get("symptom_location")) or None,
            deal_count=_to_int(get("deal_count")) or 0,
            avg_price=_to_int(get("avg_price")),
            min_price=_to_int(get("min_price")),
            max_price=_to_int(get("max_price")),
        ))
    return parts


def _parse_deals_sheet(sheet) -> list[SeedDealRow]:
    all_rows = [tuple(r) for r in sheet.iter_rows(values_only=True)]
    if not all_rows:
        return []
    header_match = _find_header_row(all_rows)
    if not header_match:
        return []
    header_idx, col_idx = header_match
    deals = []
    for row in all_rows[header_idx + 1:]:
        if not row or all(v is None for v in row):
            continue
        get = _make_getter(row, col_idx, use_positions=False)
        part_code = _clean(get("part_code"))
        if not part_code:
            continue
        deals.append(SeedDealRow(
            part_code=part_code,
            hospital=_clean(get("hospital")) or None,
            deal_date=_to_date(get("deal_date")),
            quantity=_to_int(get("quantity")) or 1,
            deal_price=_to_int(get("deal_price")),
            cost_price=_to_int(get("cost_price")),
        ))
    return deals


# ── 진입점 ────────────────────────────────────────────────────────────
def parse_parts_excel(
    file_obj: io.BytesIO,
    company_override: str | None = None,
) -> SeedData:
    wb = load_workbook(file_obj, read_only=True, data_only=True)
    sheet_names = wb.sheetnames

    parts: list[SeedRow] = []
    deals: list[SeedDealRow] = []

    # 시트 이름 기반 분리 — "부품" 시트와 "납품" 시트가 둘 다 있을 때만
    # 정규화된 분리 형식으로 처리. 그 외는 모두 비정규화로 파싱.
    part_sheet_name = None
    deal_sheet_name = None
    for sn in sheet_names:
        sn_low = sn.lower()
        if part_sheet_name is None and ("부품" in sn or "마스터" in sn or "part" in sn_low):
            if "납품" not in sn and "deal" not in sn_low and "거래" not in sn:
                part_sheet_name = sn
        if deal_sheet_name is None and ("납품" in sn or "거래" in sn or "deal" in sn_low):
            deal_sheet_name = sn

    if part_sheet_name and deal_sheet_name:
        # 명확한 분리 형식 (부품 시트 + 납품 시트)
        parts = _parse_parts_sheet(wb[part_sheet_name])
        deals = _parse_deals_sheet(wb[deal_sheet_name])
    else:
        # 비정규화 — 모든 시트를 부품+납품 혼합 형식으로 파싱
        for sn in sheet_names:
            p, d = _parse_denormalized(wb[sn], company_override)
            parts.extend(p)
            deals.extend(d)

    wb.close()

    # part_code 중복 제거
    seen: set[str] = set()
    unique_parts = []
    for p in parts:
        if p.part_code not in seen:
            seen.add(p.part_code)
            unique_parts.append(p)

    return SeedData(parts=unique_parts, deals=deals)
