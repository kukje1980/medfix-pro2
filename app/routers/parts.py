import math
import io
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from app.database import get_db
from app.schemas.part import (
    PartCreate, PartUpdate, PartResponse, PartListResponse,
    PartDealCreate, PartDealResponse, SeedData,
)
import app.crud.part as crud
from app.utils.excel_parser import parse_parts_excel

router = APIRouter(prefix="/parts", tags=["부품관리"])


@router.get("/tree")
def get_tree(db: Session = Depends(get_db)):
    return crud.get_tree(db)


@router.get("", response_model=PartListResponse)
def list_parts(
    page: int = Query(1, ge=1),
    size: int = Query(50, ge=1, le=200),
    search: str | None = None,
    company: str | None = None,
    category: str | None = None,
    model: str | None = None,
    sort_by: str = Query("deal_count", regex="^(deal_count|part_code|part_name)$"),
    db: Session = Depends(get_db),
):
    items, total = crud.get_parts(db, page=page, size=size, search=search, company=company, category=category, model=model, sort_by=sort_by)
    data = [{c.name: getattr(p, c.name) for c in p.__table__.columns} for p in items]
    return {
        "items": data,
        "total": total,
        "page": page,
        "size": size,
        "pages": math.ceil(total / size) if total else 1,
    }


@router.post("", response_model=PartResponse, status_code=201)
def create_part(data: PartCreate, db: Session = Depends(get_db)):
    obj = crud.create_part(db, data)
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}


@router.get("/{part_id}", response_model=PartResponse)
def get_part(part_id: int, db: Session = Depends(get_db)):
    obj = crud.get_part(db, part_id)
    if not obj:
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}


@router.put("/{part_id}", response_model=PartResponse)
def update_part(part_id: int, data: PartUpdate, db: Session = Depends(get_db)):
    obj = crud.update_part(db, part_id, data)
    if not obj:
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")
    return {c.name: getattr(obj, c.name) for c in obj.__table__.columns}


@router.delete("/{part_id}")
def delete_part(part_id: int, db: Session = Depends(get_db)):
    if not crud.delete_part(db, part_id):
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")
    return {"message": "삭제되었습니다."}


@router.get("/{part_id}/deals")
def get_deals(part_id: int, hospital: str | None = None, db: Session = Depends(get_db)):
    obj = crud.get_part(db, part_id)
    if not obj:
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")
    deals = crud.get_deals(db, part_id, hospital=hospital)
    return [{c.name: getattr(d, c.name) for c in d.__table__.columns} for d in deals]


@router.post("/{part_id}/deals", response_model=PartDealResponse, status_code=201)
def add_deal(part_id: int, data: PartDealCreate, db: Session = Depends(get_db)):
    deal = crud.add_deal(db, part_id, data)
    if not deal:
        raise HTTPException(status_code=404, detail="부품을 찾을 수 없습니다.")
    return {c.name: getattr(deal, c.name) for c in deal.__table__.columns}


@router.delete("/deals/{deal_id}")
def delete_deal(deal_id: int, db: Session = Depends(get_db)):
    if not crud.delete_deal(db, deal_id):
        raise HTTPException(status_code=404, detail="거래 내역을 찾을 수 없습니다.")
    return {"message": "삭제되었습니다."}


@router.post("/seed")
def seed(data: SeedData, db: Session = Depends(get_db)):
    result = crud.seed_parts(db, data)
    return result


@router.post("/seed-excel")
async def seed_excel(
    file: UploadFile = File(...),
    company: str | None = None,
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="xlsx 또는 xls 파일만 업로드 가능합니다.")
    content = await file.read()
    try:
        seed_data = parse_parts_excel(
            io.BytesIO(content),
            company_override=company or None,
        )
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Excel 파싱 오류: {str(e)}")

    # 인식된 부품이 0개면 진단 정보 반환
    if len(seed_data.parts) == 0 and len(seed_data.deals) == 0:
        # 시트 구조를 짧게 분석
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet_info = []
            for sn in wb.sheetnames[:3]:
                sh = wb[sn]
                first_rows = []
                for i, row in enumerate(sh.iter_rows(values_only=True)):
                    if i >= 3:
                        break
                    first_rows.append([str(c)[:20] if c is not None else "" for c in row[:6]])
                sheet_info.append({"name": sn, "first_rows": first_rows})
            wb.close()
        except Exception:
            sheet_info = []
        raise HTTPException(
            status_code=422,
            detail={
                "message": "인식된 부품이 없습니다. Excel 형식을 확인하세요.",
                "expected": "C열=품목코드, D열=품명 (또는 헤더에 '품목코드/품명' 포함)",
                "sheets": sheet_info,
            },
        )

    result = crud.seed_parts(db, seed_data)
    return {**result, "parsed_parts": len(seed_data.parts), "parsed_deals": len(seed_data.deals)}
